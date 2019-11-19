import tkinter
from tkinter import ttk,filedialog,messagebox,scrolledtext
from tkinter import *
import openpyxl
import threading
import time
import serial
import serial.tools.list_ports
import re
from code_dic import MY_CODE

class MY_GUI():

    #构造函数
    def __init__(self,name):
        self.init_window_name=name
    #窗口控件设置初始化
    def set_init_window(self):
        self.init_window_name.title('串口控制')
        self.init_window_name.geometry('1168x631+20+10')
        self.init_window_name['bg']='pink'
        self.init_window_name.attributes('-alpha',1)

        #按钮
        self.file_choose_button=Button(self.init_window_name,text='选择文件',bg='lightblue',width=10,command=self.thread_file)
        self.file_choose_button.place(x=40,y=170)
        self.conduct_button=Button(self.init_window_name,text='执行',bg='lightblue',width=10,command=self.com_output)
        self.conduct_button.place(x=40,y=455)
        self.log_save_button=Button(self.init_window_name,text='保存日志',bg='lightblue',width=10,command=self.thread_save)
        self.log_save_button.place(x=335,y=455)
        self.clear_button=Button(self.init_window_name,text='清空',bg='lightblue',width=10,command=self.thread_clear)
        self.clear_button.place(x=460,y=455)
        #串口选择框架
        self.com_choose_frame=Frame(self.init_window_name)
        self.com_choose_frame.place(x=20,y=12)
        #串口选择框架内部标签
        self.com_label=Label(self.com_choose_frame,text='COMx: ')
        self.com_label.grid(row=0,column=0,sticky=E)
        self.baudrate_label=Label(self.com_choose_frame,text='Baudrate: ')
        self.baudrate_label.grid(row=1,column=0,sticky=E,pady=10)
        #串口框架内部下拉选项框
        self.com_choose=StringVar()
        self.com_choose_combo=ttk.Combobox(self.com_choose_frame,width=30,textvariable=self.com_choose)
        self.com_choose_combo['state']='readonly'
        self.com_choose_combo.grid(row=0,column=1,padx=15)
        self.com_choose_combo['values']=self.com_name_get()
        #波特率选项框
        self.baudrate_value=StringVar(value='9600')
        self.baudrate_choose_combo=ttk.Combobox(self.com_choose_frame,width=30,textvariable=self.baudrate_value)
        self.baudrate_choose_combo['values']=('9600','115200')
        self.baudrate_choose_combo['state']='readonly'
        self.baudrate_choose_combo.grid(row=1,column=1,padx=15)
        #串口框架内部按钮
        self.connect_button=Button(self.com_choose_frame,text='连接',bg='lightblue',width=10,command=self.com_connect)
        self.connect_button.grid(row=0,column=2,padx=15)
        self.cancel_button=Button(self.com_choose_frame,text='取消',bg='lightblue',width=10,command=self.com_cancel)
        self.cancel_button.grid(row=1,column=2,padx=15)

        #处理结果显示滚动文本框
        self.result_text=scrolledtext.ScrolledText(self.init_window_name,width=77,height=42)
        self.result_text.place(x=600,y=50)

        #其他文本框
        self.file_path_text=Text(self.init_window_name,width=57,height=1)
        self.file_path_text.place(x=150,y=175)
        self.com_log_text=Text(self.com_choose_frame,width=78,height=5)
        self.com_log_text.grid(row =2,column =0,columnspan=3,pady=5)
        self.com_log_text.insert(END,'此处显示串口工作信息'+'\n')

        #标签
        self.result_data_label=Label(self.init_window_name,text='输出结果')
        self.result_data_label.place(x=600,y=15)
        self.num_input_label=Label(self.init_window_name,text='输入编号： ')
        self.num_input_label.place(x=150,y=460)

        #输入编号框
        self.input_num=StringVar()
        self.input_num_entry=Entry(self.init_window_name,textvariable=self.input_num,width=10)
        self.input_num_entry.place(x=215,y=460)

        #代码解析后进行显示
        self.code_frame=Frame(self.init_window_name,width=78,height=29,bg='white')
        self.code_frame.place(x=20,y=210)
        #解析后的代码放在表格内显示
        self.code_tree=ttk.Treeview(self.code_frame,show='headings',height=10,columns=('0','1','2','3','4'))
        self.code_bar=ttk.Scrollbar(self.code_frame,orient=VERTICAL,command=self.code_tree.yview)
        self.code_tree.configure(yscrollcommand=self.code_bar.set)
        self.code_tree.grid(row=0,column=0,sticky=NSEW)
        self.code_bar.grid(row=0,column=1,sticky=NS)
        self.code_tree.column('0',width=30)
        self.code_tree.column('1',width=250)
        self.code_tree.column('2',width=50)
        self.code_tree.column('3',width=100)
        self.code_tree.column('4',width=100)
        self.code_tree.heading('0',text='序号')
        self.code_tree.heading('1',text='命令')
        self.code_tree.heading('2',text='状态')
        self.code_tree.heading('3',text='失败则执行次数')
        self.code_tree.heading('4',text='再失败则跳转至')

        #执行结果显示frame
        self.result_frame=Frame(self.init_window_name,width=78,height=15,bg='white')
        self.result_frame.place(x=20,y=500)
        self.result_tree=ttk.Treeview(self.result_frame,show='headings',height=4,columns=('0','1','2'))
        self.result_bar=ttk.Scrollbar(self.result_frame,orient=VERTICAL,command=self.result_tree.yview)
        self.result_tree.configure(yscrollcommand=self.result_bar.set)
        self.result_tree.grid(row=0,column=0,sticky=NSEW)
        self.result_bar.grid(row=0,column=1,sticky=NS)
        self.result_tree.column('0',width=30)
        self.result_tree.column('1',width=80)
        self.result_tree.column('2',width=420)
        self.result_tree.heading('0',text='编号')
        self.result_tree.heading('1',text='是否成功')
        self.result_tree.heading('2',text='从哪句指令开始失败')

    #自动获取当前连接的串口名
    def com_name_get(self):
        self.port_list=list(serial.tools.list_ports.comports())
        self.com_port_names=[]
        self.pattern=re.compile(r'[(](.*?)[)]',re.S)
        if len(self.port_list)>0:
            for i in range(len(self.port_list)):
                self.com_name=re.findall(self.pattern,str(self.port_list[i]))
                self.com_port_names.append(self.com_name)
        return self.com_port_names

    #连接按键的执行内容
    def com_connect(self):
        self.result_text.insert(END,'请连接串口设备'+'\n')
        self.ser_name=str(self.com_choose.get())
        self.ser_baudrate=str(self.baudrate_value.get())
        try:
            self.ser=serial.Serial(self.ser_name)
            self.ser.baudrate=self.ser_baudrate
            self.ser.timeout=0.5
            self.com_log_text.insert(END,time.ctime(time.time())+'\t\t'+'串口成功打开'+'\n')
            self.com_log_text.see(tkinter.END)
            self.com_log_text.update()
            while True:
                newline=self.ser.readline()#字节类型
                self.result_text.insert(END,newline)
                self.result_text.see(tkinter.END)
                self.result_text.update()
        except:
            newline=time.ctime(time.time())+'\t\t'+'串口打开故障或串口被关闭'+'\n'
            self.com_log_text.insert(END,newline)
            self.com_log_text.see(tkinter.END)
            self.com_log_text.update()

    #取消按键的执行内容
    def com_cancel(self):
        self.result_text.delete('1.0','end')
        try:
            self.ser.close()
            newline=time.ctime(time.time())+'\t\t'+'串口被关闭'+'\n'
        except:
            newline=time.ctime(time.time())+'\t\t'+'串口未打开'+'\n'
        self.com_log_text.insert(END,newline)
        self.com_log_text.see(tkinter.END)
        self.com_log_text.update()

    #执行按键的执行内容
    def com_output(self):
        autodocument=open('autoresult.txt','a')
        self.code_tree_items=self.code_tree.get_children()
        for i in range(self.codeline_counter):
            self.code_tree.set(self.code_tree_items[i],column=2,value=' ')
        self.code_result_sheet=['failed'for i in range(self.codeline_counter)]
        self.functionname=[]
        if self.input_num_entry.get():
            self.testnum=self.input_num_entry.get()
        else:
            self.testnum='未标'
        newline='编号: '+'\t'+str(self.testnum)
        print(newline,file=autodocument)
        try:
            for i in range(self.codeline_counter):
                function=str(self.code_sheet[i][1])
                self.functionandparam=re.split(',',function)                              #用，将函数名与参数分开
                functionname=self.functionandparam[0]                                     #第一个元素取为函数名，后面的都是参数
                self.functionname.append(functionname)
                my_code=MY_CODE(self.ser,self.result_text,self.functionandparam)
                result=my_code.function_choose_do(functionname)
                if result ==1:
                    self.code_tree.set(self.code_tree_items[i],column=2,value='ok')
                    self.code_result_sheet[i]='ok'
                else:
                    self.code_tree.set(self.code_tree_items[i],column=2,value='failed')
                    self.code_result_sheet[i]='failed'
                newline=str(self.code_sheet[i][1])+':'+'\t'+str(self.code_result_sheet[i])
                print(newline,file=autodocument)
        except:
            tkinter.messagebox.showerror('执行异常','代码执行异常，请仔细检查代码及格式！')
        failnum = 0
        for code_result in self.code_result_sheet:
            if code_result == 'ok':
                failnum += 1
            else:
                break
        if failnum < self.codeline_counter :
            self.isok = '失败'
            self.failcode=str(self.code_sheet[failnum][1])
        else:
            self.isok='成功'
            self.failcode = ' '
        self.result_tree.insert('','end',values=[self.testnum,self.isok,self.failcode])


    #新建线程，负责选择代码文件、保存代码执行结果和清空代码表格

    #新建选择文件线程
    def thread_file(self):
        thisthread=threading.Thread(target=self.file_choose)
        thisthread.start()

    #选择文件打开，并在界面中显示
    def file_choose(self):
        self.codeline_counter=0
        self.root=Tk()
        self.root.withdraw()
        file_path=filedialog.askopenfilename()
        self.file_path_text.insert(END,file_path)
        wb=openpyxl.load_workbook(file_path)
        sheet=wb['Sheet1']
        self.code_sheet=[[0 for i in range(6)]for j in range(20)]
        for i in range(20):
            if sheet.cell(row=i+2,column=1).value:
                self.codeline_counter +=1
                self.code_context=[]
                for j in range(6):
                    self.code_sheet[i][j]=sheet.cell(row=i+2,column=j+1).value
                    self.code_context.append(self.code_sheet[i][j])
                self.code_tree.insert('',i,values=self.code_context)  #code_tree用来将代码规则的显示，code_sheet是一个数组，便于取值运算

    #新建线程保存执行结果
    def thread_save(self):
        thisthread=threading.Thread(target=self.code_log_save)
        thisthread.start()

    #保存代码执行结果日志
    def code_log_save(self):
        document=open('result.txt','a')#a表示写入时不将原有内容清除
        try:
            firstline='编号： '+str(self.testnum)
        except:
            firstline='编号： '+'未标注'
        print(firstline,file=document)
        secondline='指令'.rjust(10,' ')+'\t'+'执行结果'
        print(secondline,file=document)
        try:
            for i in range(self.codeline_counter):
                print(str(self.functionname[i]).rjust(12,' ')+': '+'\t'+str(self.code_result_sheet[i]),file=document)
            tkinter.messagebox.showinfo('保存成功','Saved successfully!')
        except:
            thirdline='\t\t'+'代码未执行'
            print(thirdline,file=document)
            tkinter.messagebox.showinfo('异常','没有选择文件或代码执行异常')

    #新建线程清空所选文件以备重新选择
    def thread_clear(self):
        thisthread=threading.Thread(target=self.file_clear)
        thisthread.start()

    #删除所选文件，清空解析后的代码表格
    def file_clear(self):
        self.file_path_text.delete('1.0','end')
        self.input_num.set(' ')
        code_items=self.code_tree.get_children()
        for item in code_items:
            self.code_tree.delete(item)
        result_items=self.result_tree.get_children()
        for item in result_items:
            self.result_text.delete(item)

#主线程
def start():
    init_window=Tk()
    my_window=MY_GUI(init_window)
    my_window.set_init_window()
    d=open('autoresult.txt','wb+')
    d.truncate()
    f=open('result.txt','wb+')
    f.truncate()                                                               #每次运行时将文件清空，每一次保存日志时，文件不清空
    init_window.mainloop()

start()

